#
# File: xlsrw_oop.py
# Subject: Compare two Excel files which include invoice information and general ledger for monthly
#          accounting purpose
# Brief: Entry of command line invoice matching tool
# Coder: alfan-ntu
# Created Date: 2020/10/2
# Revision:
#   1. 2020/10/2: v. 0.1 1st creation
#   2. 2020/10/30: v. 0.2
#           - rearrange utility functions by moving some widgets to utility.py
#           - modify Invoice Details to include a matched mark
#   3. 2020/11/1: v. 0.3
#           - split the Excel processing to pre-pro and record-matching
#   4. 2020/11/3: v. 0.4
#           - progress bar displayed when traversing the source invoice details data
#   5. 2021/4/6: v. 1.1
#           - support both CLI and GUI
#
# ToDo's:
#   1) Add invoice date range
#   2) Apply filter to the result Excel file; and freeze the top row of the result Excel file
#
# Note:
#   1. xlrd can extract data from Excel files of format, .xls or .xlsx
#   2. xlwt can generate spreadsheet file compatible with .xls (MS Excel 97/2000/XP/200) BUT NOT .xlsx
#   3. openpyxl can read/write Excel 2010 xlsx/xlsm/xltx/xltm files, BUT NOT .xls
#   Both xlrd/xlwt/xlutils and openpyxl packages are required to process the input
#   invoice files and general ledger files since Invoice Details is .xls while General Ledger
#   is .xlsx. Pandas might be a flexible and more versatile alternative.
#
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import xlsxwriter
import pdb
import xlrd
from xlutils.copy import copy as xlutils_copy
import progressbar

import constant
import class_transaction
import utility
import logging
import class_opts


#
# Function: match_row(sourceRow, targetWs)
# Subject: check if any record found in targetWs that
#   - buyerName,
#   - invoice status(開立、作廢),
#   - invoice date
#   - invoice amount
#   in sourceRow match the record in targetWs
# Input :
#       - sourceRow: the whole row sequence in the source invoice records Excel
#       - targetWs: the target worksheet which target general ledger stores in
# Output :
#       - a tuple
#
def match_row(sourceRow, targetWs):
    # pdb.set_trace()
    invoice_status = sourceRow[constant.COL_INVOICE_STATUS].value
    matchRow_in_targetWs = 0
    if invoice_status == "作廢":
        # print(sourceRow[6], "/", invoice_status)
        # print(sourceRow[constant.COL_INVOICE_NO].value, " is void")
        # return immediately if the invoice is void
        return False, matchRow_in_targetWs
    #
    # Continue traversing target worksheet only if the invoice in the sourceRow is a valid one
    # print(sourceRow[constant.COL_INVOICE_NO].value, " is valid")
    # for jt in (1, targetWs.nrows):
    ex_rate = utility.find_currency_exchange_rate(sourceRow)
    if ex_rate == 1.0:
        pass
    else:
        print(sourceRow[constant.COL_INVOICE_NO].value, " is a transaction in USD$, exchange rate:", str(ex_rate))
    return True, 20


#
# Filter general ledger file and leave Account Receivables only in external sales in the target Excel file
#
def preproc_general_ledger(gl_excel, ext_sales_excel, GUI_caller):
    # check caller type
    # pdb.set_trace()
    if GUI_caller:
        print("preproc_general_ledger is called from GUI")
        print("General ledger " + gl_excel)
    wb_src= openpyxl.load_workbook(gl_excel, read_only=True)
    ws_name = wb_src.sheetnames[0]
    ws_src = wb_src[ws_name]

    wb_tgt = openpyxl.Workbook()
    ws_tgt = wb_tgt.create_sheet("Sheet0", 0)
    ws_tgt.sheet_format.defaultColWidth = 12
    ws_tgt.column_dimensions["C"].width = 16
    ws_tgt.column_dimensions["O"].width = 28
    header_row = True
    for r in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row):
        if header_row:
            header_row = False
            ws_tgt.append(cell.value for cell in r)
            continue
        # only transactions with voucher type = "F" and account description includes "Accounts Receivable"
        # are required
        voucher_type = r[constant.COL_GL_VOUCHER_TYPE].value
        accnt_desc = r[constant.COL_GL_ACCOUNT_DESCRIPTION].value
        idx_accnt_desc = accnt_desc.find(constant.TARGET_ACCOUNT_IN_GL)
        if idx_accnt_desc > 0 and voucher_type == "F":
            ws_tgt.append(cell.value for cell in r)

    wb_src.close()
    num_of_col = ws_tgt.max_column
    num_of_row = ws_tgt.max_row
    ws_tgt.auto_filter.ref = "A1:Z1"
    # Just an experiment
    # ws_tgt.auto_filter.add_filter_column(12, ["TW72"])
    ws_tgt.insert_cols(4, 3)
    ws_tgt.cell(row=1, column=4, value="統一發票號碼")
    ws_tgt.cell(row=1, column=5, value="發票稅後\t台幣總金額\t(美金報價)")
    ws_tgt.cell(row=1, column=6, value="比對")
    ws_tgt.column_dimensions["D"].width = 20
    ws_tgt.column_dimensions["E"].alignment = Alignment(wrapText=True)
    for r in range(1, ws_tgt.max_row+1):
        for c in range(1, ws_tgt.max_column+1):
            ws_tgt.cell(row=r, column=c).font = Font(name="Calibri")
    wb_tgt.save(ext_sales_excel)
    wb_tgt.close()
    # Notify GUI that general ledger pre-process is done
    if GUI_caller:
        GUI_caller.gl_prep_done_ev.set()

    return True


#
# match invoice details(sourceWb, source workbook) to external sales
# records(processed General ledger)
#
def match_invoice_and_external_sales(invoice_excel, ext_sales_excel, GUI_caller):
    # Open source invoice details Excel file, which is of .xls format
    sourceWb = xlrd.open_workbook(invoice_excel, formatting_info=True)
    sheetName = "Sheet0"
    sourceWs = sourceWb.sheet_by_name(sheetName)
    sourceWb_temp = xlutils_copy(sourceWb)
    sourceWs_temp = sourceWb_temp.get_sheet(0)
    # check caller type
    if GUI_caller:
        print("match_invoice_and_external_sales is called from GUI")

    #
    # openpyxl to read external sales Excel file in order to read/modify/write .xlsx files
    # External sales Excel file was created in the stage 總帳前處理
    #
    # targetWb = openpyxl.load_workbook(ext_sales_excel)
    ext_sales_wb = openpyxl.load_workbook(ext_sales_excel)
    # 0-based index, index of worksheet #1 is 0

    ext_sales_ws_name = ext_sales_wb.sheetnames[0]
    ext_sales_ws = ext_sales_wb[ext_sales_ws_name]
    #
    # Traverse the source invoice records
    # pdb.set_trace()
    number_of_matched_found = 0

    sourceWs_temp.write(0,constant.COL_INVOICE_CHECKED, "發票配對")
    bar = progressbar.ProgressBar(maxval=100, widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
    bar.start()
    for js in range(1, sourceWs.nrows):
        p = (js/sourceWs.nrows) * 100
        bar.update(p)
        invoice_status = sourceWs.cell_value(js, constant.COL_INVOICE_STATUS)
        if invoice_status == "作廢":
            sourceWs_temp.write(js, constant.COL_INVOICE_CHECKED, "作廢")
            continue
        invoice_number = sourceWs.cell_value(js, constant.COL_INVOICE_NO)
        buyer_name = sourceWs.cell_value(js, constant.COL_INVOICE_BUYER)
        invoice_date = sourceWs.cell_value(js, constant.COL_INVOICE_DATE)
        amount_nt_str = sourceWs.cell_value(js, constant.COL_INVOICE_TOTAL)
        invoice_amount_nt = utility.comma_separated_amount_to_float(amount_nt_str)
        # determine if it is a USD transaction, extract the exchange rate if
        # it is a USD transaction
        if utility.is_source_a_usd_transaction(sourceWs.row(js)):
            function_currency = constant.FUNCTION_CURRENCY_USD
            exchange_rate = utility.find_currency_exchange_rate(sourceWs.row(js))
        else:
            function_currency = constant.FUNCTION_CURRENCY_NTD
            exchange_rate = 1.00
        source = constant.DATA_SOURCE_INVOICE_DETAIL
        source_transaction = class_transaction.Transaction(invoice_number,
                                                          buyer_name,
                                                          invoice_date,
                                                          invoice_amount_nt,
                                                          0.0,
                                                          function_currency,
                                                          exchange_rate,
                                                          source)
        # call the class method to display the object contents
        source_transaction.display_transaction()
        # traverse the target worksheet and identify the record correspondent to the
        # source transaction
        match_found = False

        # for jt in range(2, targetWs.max_row+1):
        for jt in range(2, ext_sales_ws.max_row+1):
            if not utility.is_target_account_receivable(ext_sales_ws[jt]):
                continue
            invoice_number = ext_sales_ws.cell(row=jt, column=constant.COL_ES_INVOICE_NO+1).value
            buyer_name = ext_sales_ws.cell(row=jt, column=constant.COL_ES_TEXT+1).value
            invoice_date = ext_sales_ws.cell(row=jt, column=constant.COL_ES_INVOICE_DATE+1).value
            invoice_amount_nt = ext_sales_ws.cell(row=jt, column=constant.COL_ES_AMOUNT+1).value
            if utility.is_target_a_usd_transaction(ext_sales_ws[jt]):
                function_currency = constant.FUNCTION_CURRENCY_USD
                exchange_rate = ext_sales_ws.cell(row=jt, column=constant.COL_ES_EXCHANGE_RATE+1).value
                invoice_amount_us = invoice_amount_nt / exchange_rate
            else:
                function_currency = constant.FUNCTION_CURRENCY_NTD
                exchange_rate = 1.0
                invoice_amount_us = 0.0
            source = constant.DATA_SOURCE_GENERAL_LEDGER
            target_transaction = class_transaction.Transaction(invoice_number,
                                                               buyer_name,
                                                               invoice_date,
                                                               invoice_amount_nt,
                                                               invoice_amount_us,
                                                               function_currency,
                                                               exchange_rate,
                                                               source)
            if source_transaction.match_transaction(target_transaction):
                match_found = True
                logging.info(">>>>>>>>>>>>>> 找到匹配交易紀錄 <<<<<<<<<<<<<<<")
                number_of_matched_found += 1
                logging.info("已匹配交易數量: %d", number_of_matched_found)
                target_transaction.display_transaction()
                logging.info("==========================================================")
                sourceWs_temp.write(js, constant.COL_INVOICE_CHECKED, "是")
                ext_sales_ws.cell(row=jt,
                                  column=constant.COL_ES_UNIFIED_INVOICE_NO+1,
                                  value=source_transaction.invoice_number)
                # if source_transaction.exchange_rate != 1.0:
                #     ext_sales_ws.cell(row=jt,
                #                       column=constant.COL_ES_UINV_AMT+1,
                #                       value=source_transaction.invoice_amount_NT)
                #     ext_sales_ws.cell(row=jt, column=constant.COL_ES_UINV_AMT+1).number_format='"$"#,##0_-'
                ext_sales_ws.cell(row=jt,
                                  column=constant.COL_ES_UINV_AMT+1,
                                  value=source_transaction.invoice_amount_NT)
                ext_sales_ws.cell(row=jt, column=constant.COL_ES_UINV_AMT+1).number_format='"$"#,##0_-'
                ext_sales_ws.cell(row=jt,
                                  column=constant.COL_ES_INVOICE_MATCHED+1,
                                  value="配對")

        if jt == ext_sales_ws.max_row and match_found is False:
            logging.info(">>>>>>>>>>>>>> 無法找到匹配交易紀錄 <<<<<<<<<<<<<<<, ext_sales_ws.max_row %s", ext_sales_ws.max_row)
            logging.info("==========================================================")
            sourceWs_temp.write(js, constant.COL_INVOICE_CHECKED, "否")
    bar.finish()
    if GUI_caller:
        GUI_caller.print_log("3. 原始發票資料檔比對完成，比對結果註記在 %s 的'發票配對'欄位" % invoice_excel)
    else:
        print("3. 原始發票資料檔比對完成，比對結果註記在 %s 的'發票配對'欄位" % invoice_excel)
    sourceWb_temp.save(invoice_excel)
    if GUI_caller:
        GUI_caller.print_log("4. 總帳濾出應收帳款資料，儲存於 %s" % ext_sales_excel)
    else:
        print("4. 總帳濾出應收帳款資料，儲存於 %s" % ext_sales_excel)
    ext_sales_wb.save(ext_sales_excel)
    return True


#
# main entry of Command Line Executable
#
def main(argv):
    # process argv and opts
    opts_args = class_opts.Opts(argv)
    # Initialize the execution
    utility.initialization()
    #
    # Fetch target general ledger Excel file and external sales Excel file
    #
    invoice_details = opts_args.invoice_file
    general_ledger = opts_args.ledger_file
    external_sales = opts_args.sales_file
    #
    # fetch invoice duration information
    #
    if opts_args.begin_date != "":
        print("對帳起始日期: ", opts_args.begin_date.strftime("%Y/%m/%d"))
    if opts_args.end_date != "":
        print("對帳截止日期: ", opts_args.end_date.strftime("%Y/%m/%d"))

    print("1. 進行總帳前處理")
    preproc_general_ledger(general_ledger, external_sales, None)
    print("2. 進行原始發票資料檔比對")
    match_invoice_and_external_sales(invoice_details, external_sales, None)


def generate_excel(spread_sheet):
    workbook = xlsxwriter.Workbook("Customs_bill_records.xlsx")
    worksheet = workbook.add_worksheet("2nd_sheet")
    tax_bill = ""
    customs_bill = ""
    tax_ID = ""
    tax_amount = ""
    row = 0
    col = 0
    for tax_bill, customs_bill, tax_ID, tax_amount in spread_sheet:
        worksheet.write(row, col, tax_bill)
        worksheet.write(row, col+1, customs_bill)
        worksheet.write(row, col+2, tax_ID)
        worksheet.write(row, col+3, tax_amount)
        row += 1

    workbook.close()


def get_input_file():
    input_file_name = "201903.txt"
    hInputFile = open(input_file_name, "r", encoding="utf-8")
    return hInputFile


if __name__ == "__main__":
    main(sys.argv[0:])
